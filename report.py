"""
Report generation: an HTML preview (for on-screen display) and a formatted
.xlsx workbook (for download), both built from the same live state — so a
change to categories/items/columns is reflected in both immediately.

Supplier-matrix layout matches the org's original template exactly: each
supplier gets three columns — Harga/Unit (RM), Jumlah (RM), and Status —
plus one shared "Catatan" column per item at the very end of the row (not
per supplier). Harga/Unit and Status both accept free text as well as
numbers/codes (e.g. Kewangan typing "TT" for Tiada Tawaran, or Teknikal
typing a custom status beyond M/TM/TT).
"""

import io
import html as htmlmod

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import state as S


def _esc(v):
    return htmlmod.escape(str(v)) if v is not None else ""


def _jumlah_value(bid, qty):
    """What the Jumlah/K cell shows: the computed total (numeric) if a real
    price was given, else whatever free-text override was typed into
    Harga/Unit, else the template's standard 'TT' (Tiada Tawaran) fallback."""
    up = bid.get("unit_price")
    if up is not None:
        return qty * up, True
    if bid.get("price_text"):
        return bid["price_text"], False
    return "TT", False


# --------------------------------------------------------------------------
# HTML preview
# --------------------------------------------------------------------------

def build_report_html(state: dict) -> str:
    suppliers = state["suppliers"]
    cats = S.sorted_categories(state)
    cols = S.sorted_columns(state)
    cfg = state["supplier_matrix"]
    sub_count = max((1 if cfg["show_unit_price"] else 0) + (1 if cfg["show_total"] else 0)
                     + (1 if cfg["show_status"] else 0), 1)

    parts = []
    parts.append('<div style="text-align:center;font-size:12px;letter-spacing:.05em;'
                  'text-transform:uppercase;color:#4a5b6b">LAMPIRAN</div>')
    title = _esc(state["meta"]["title"])
    ref = _esc(state["meta"]["ref"])
    parts.append(f'<div style="text-align:center;font-weight:800;font-size:15px;margin:4px 0 14px;'
                  f'color:#142a3a">{title}{" : " + ref if ref else ""}</div>')

    parts.append('<table style="border-collapse:collapse;width:100%;font-size:11px" border="1">')
    parts.append("<thead>")
    parts.append('<tr style="background:#1f3d52;color:#fff">')
    parts.append('<th rowspan="2" style="padding:5px;min-width:44px">Bil.</th>')
    parts.append('<th rowspan="2" style="padding:5px;min-width:200px">Perkara</th>')
    total_cols = 2
    for col in cols:
        if col["type"] == "supplier-matrix":
            for i, s in enumerate(suppliers):
                parts.append(f'<th colspan="{sub_count}" style="padding:5px">{i+1}/{len(suppliers)}<br>'
                              f'<span style="font-weight:400">{_esc(s["name"])}</span></th>')
                total_cols += sub_count
        else:
            parts.append(f'<th rowspan="2" style="padding:5px">{_esc(col["label"])}</th>')
            total_cols += 1
    parts.append("</tr><tr style=\"background:#1f3d52;color:#fff\">")
    for col in cols:
        if col["type"] == "supplier-matrix":
            for _ in suppliers:
                if cfg["show_unit_price"]:
                    parts.append('<th style="padding:4px;font-size:10px">Harga/Unit (RM)</th>')
                if cfg["show_total"]:
                    parts.append('<th style="padding:4px;font-size:10px">Jumlah (RM)</th>')
                if cfg["show_status"]:
                    parts.append('<th style="padding:4px;font-size:10px">Status</th>')
    parts.append("</tr></thead><tbody>")

    for cat in cats:
        items = S.items_for_category(state, cat["id"])
        is_parent = bool(S.subcategories_of(state, cat["id"]))
        cat_num = S.category_full_number(state, cat["id"])
        parts.append(f'<tr style="background:#f2e6d3;font-weight:700;color:#142a3a">'
                      f'<td colspan="{total_cols}" style="padding:5px">{cat_num}. {_esc(cat["name"])}</td></tr>')

        subtotals = {}
        supplier_subtotals = [{"unit": 0.0, "total": 0.0} for _ in suppliers]
        qty_col = next((c for c in state["columns"] if c.get("tag") == "quantity"), None)

        for idx, item in enumerate(items):
            parts.append(f'<tr><td style="padding:4px;text-align:center">{cat_num}.{idx+1}</td>'
                          f'<td style="padding:4px;white-space:pre-line">{_esc(item["perkara"])}</td>')
            for col in cols:
                if col["type"] == "supplier-matrix":
                    qty = float(S.get_field(item, qty_col["id"]) or 0) if qty_col else 1
                    for si, s in enumerate(suppliers):
                        bid = S.get_bid(state, item["id"], s["id"])
                        up = bid.get("unit_price")
                        if up is not None:
                            supplier_subtotals[si]["unit"] += up
                            supplier_subtotals[si]["total"] += qty * up
                        st_def = S.status_def(state, bid.get("status"))
                        fill_style = ""
                        if st_def:
                            colors = S.STATUS_COLORS.get(st_def["color"], S.STATUS_COLORS["neutral"])
                            fill_style = f'background:{colors["bg"]};color:{colors["fg"]}'
                        if cfg["show_unit_price"]:
                            val = _esc(S.price_display(bid))
                            parts.append(f'<td style="padding:4px;text-align:right">{val}</td>')
                        if cfg["show_total"]:
                            jval, is_num = _jumlah_value(bid, qty)
                            val = S.fmt_money(jval) if is_num else _esc(jval)
                            parts.append(f'<td style="padding:4px;text-align:right">{val}</td>')
                        if cfg["show_status"]:
                            val = _esc(bid.get("status") or "")
                            parts.append(f'<td style="padding:4px;text-align:center;font-weight:700;{fill_style}">{val}</td>')
                else:
                    v = S.compute_column_value(item, col)
                    if col["type"] in ("number", "currency", "computed"):
                        if isinstance(v, (int, float)):
                            subtotals[col["id"]] = subtotals.get(col["id"], 0) + v
                        parts.append(f'<td style="padding:4px;text-align:right">{S.format_column_value(item, col)}</td>')
                    elif col["type"] == "supplier-pick":
                        name = S.supplier_name(state, v) if v else ""
                        parts.append(f'<td style="padding:4px;text-align:center;font-size:10px">{_esc(name)}</td>')
                    else:
                        parts.append(f'<td style="padding:4px;font-size:10px">{_esc(v or "")}</td>')
            parts.append("</tr>")

        if not items:
            if not is_parent:
                parts.append(f'<tr><td colspan="{total_cols}" style="padding:6px;font-style:italic;color:#888">Tiada item.</td></tr>')
        else:
            parts.append('<tr style="background:#eef1f3;font-weight:700"><td colspan="2" style="padding:4px;text-align:center">Jumlah Kecil</td>')
            for col in cols:
                if col["type"] == "supplier-matrix":
                    for si, s in enumerate(suppliers):
                        if cfg["show_unit_price"]:
                            v = supplier_subtotals[si]["unit"]
                            parts.append(f'<td style="padding:4px;text-align:right">{S.fmt_money(v) if v else ""}</td>')
                        if cfg["show_total"]:
                            v = supplier_subtotals[si]["total"]
                            parts.append(f'<td style="padding:4px;text-align:right">{S.fmt_money(v) if v else ""}</td>')
                        if cfg["show_status"]:
                            parts.append("<td></td>")
                elif col["type"] in ("number", "currency", "computed"):
                    v = subtotals.get(col["id"])
                    parts.append(f'<td style="padding:4px;text-align:right">{S.fmt_money(v) if v else ""}</td>')
                else:
                    parts.append("<td></td>")
            parts.append("</tr>")

    parts.append("</tbody></table>")
    if not cats:
        parts.append('<p style="color:#4a5b6b;font-style:italic">Tiada kategori lagi.</p>')

    if cats:
        parts.append(_build_summary_html(state, cols, suppliers, cfg))

    return "".join(parts)


def _build_summary_html(state, cols, suppliers, cfg):
    summ = S.compute_summary(state)
    sub_count = max((1 if cfg["show_unit_price"] else 0) + (1 if cfg["show_total"] else 0)
                     + (1 if cfg["show_status"] else 0), 1)

    def supplier_cells(value_map):
        html = ""
        for s in suppliers:
            v = value_map.get(s["id"], 0)
            shown = False
            if cfg["show_unit_price"]:
                html += "<td></td>"
            if cfg["show_total"]:
                html += f'<td style="padding:4px;text-align:right;color:#1c2b39">{S.fmt_money(v) if v else ""}</td>'
                shown = True
            if cfg["show_status"]:
                html += "<td></td>"
            if not shown and not cfg["show_unit_price"] and not cfg["show_status"]:
                html += f'<td style="padding:4px;text-align:right;color:#1c2b39">{S.fmt_money(v) if v else ""}</td>'
        return html

    def blank_supplier_cells():
        return "<td></td>" * sub_count

    def single_value_row(label, value, place_under_esttotal=True):
        row = f'<tr style="background:#dce8f5;font-weight:700"><td colspan="2" style="padding:5px;color:#1c2b39">{_esc(label)}</td>'
        placed = False
        for col in cols:
            if col["type"] == "supplier-matrix":
                row += blank_supplier_cells()
            else:
                if place_under_esttotal and col.get("tag") == "estTotal" and not placed:
                    row += f'<td style="padding:4px;text-align:right;color:#1c2b39">{S.fmt_money(value)}</td>'
                    placed = True
                else:
                    row += "<td></td>"
        if not placed:
            row += f'<td style="padding:4px;text-align:right;color:#1c2b39">{S.fmt_money(value)}</td>'
        row += "</tr>"
        return row

    def per_supplier_row(label, value_map):
        row = f'<tr style="background:#dce8f5;font-weight:700"><td colspan="2" style="padding:5px;color:#1c2b39">{_esc(label)}</td>'
        for col in cols:
            if col["type"] == "supplier-matrix":
                row += supplier_cells(value_map)
            else:
                row += "<td></td>"
        row += "</tr>"
        return row

    parts = ['<table style="border-collapse:collapse;width:100%;font-size:11px;margin-top:2px" border="1"><tbody>']
    parts.append(single_value_row("HARGA ANGGARAN JABATAN (RM)", summ["dept_total"]))
    parts.append(per_supplier_row("HARGA TAWARAN DARI PENYEBUT HARGA (RM)", summ["supplier_totals"]))
    if summ["has_pick_column"]:
        parts.append(per_supplier_row("HARGA YANG DICADANGKAN (RM)", summ["supplier_recommended"]))
    else:
        parts.append(
            f'<tr style="background:#faeadc"><td colspan="2" style="padding:5px;color:#b3541e">HARGA YANG DICADANGKAN (RM)</td>'
            f'<td colspan="{max(len(cols)-1,1)}" style="padding:5px;color:#b3541e;font-style:italic">'
            f'Tetapkan pilihan Pembekal Dipilih bagi setiap item untuk mengira baris ini.</td></tr>'
        )
    parts.append(single_value_row("JUMLAH KESELURUHAN PEROLEHAN (RM)", summ["grand_total"], place_under_esttotal=False))
    parts.append("</tbody></table>")
    return "".join(parts)


# --------------------------------------------------------------------------
# Excel export
# --------------------------------------------------------------------------

NAVY = "FF1F3D52"
GOLD = "FFF2E6D3"
SUBTOTAL_BG = "FFEEF1F3"
COLOR_HEX = {"ok": "FFE5F2E9", "bad": "FFF7E3E3", "warn": "FFFAEADC", "neutral": "FFEEEEEE"}
THIN = Side(style="thin", color="FF9AA7B0")
BORDER = Border(top=THIN, left=THIN, bottom=THIN, right=THIN)


def _header_cell(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, color="FFFFFFFF", size=9)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _scoped_supplier_fields(cfg, scope):
    """Which of the 3 supplier sub-columns (Harga/Unit, Jumlah, Status) to
    show for this export scope."""
    if scope == "kewangan":
        return {"unit_price": cfg["show_unit_price"], "total": cfg["show_total"], "status": False}
    if scope == "technical":
        return {"unit_price": False, "total": False, "status": cfg["show_status"]}
    return {"unit_price": cfg["show_unit_price"], "total": cfg["show_total"], "status": cfg["show_status"]}


def _column_included(col, scope):
    """Whether a non-supplier-matrix editable column belongs in this export scope."""
    if scope == "full":
        return True
    team = col.get("team", "shared")
    return team in (scope, "shared")


def build_excel(state: dict, scope: str = "full") -> bytes:
    """scope: 'full' (merged), 'kewangan' (financial-only), or 'technical' (technical-only)."""
    suppliers = state["suppliers"]
    cats = S.sorted_categories(state)
    all_cols = S.sorted_columns(state)
    cols = [c for c in all_cols if c["type"] == "supplier-matrix" or _column_included(c, scope)]
    cfg = state["supplier_matrix"]
    fields = _scoped_supplier_fields(cfg, scope)
    sub_count = max((1 if fields["unit_price"] else 0) + (1 if fields["total"] else 0)
                     + (1 if fields["status"] else 0), 1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan"

    total_cols = 2
    for c in cols:
        total_cols += len(suppliers) * sub_count if c["type"] == "supplier-matrix" else 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c1 = ws.cell(row=1, column=1, value="LAMPIRAN")
    c1.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ref = state["meta"].get("ref", "")
    c2 = ws.cell(row=2, column=1, value=state["meta"]["title"] + (f' : {ref}' if ref else ""))
    c2.font = Font(bold=True, size=12)
    c2.alignment = Alignment(horizontal="center")

    hr1, hr2 = 4, 5
    _header_cell(ws, hr1, 1, "Bil."); ws.merge_cells(start_row=hr1, start_column=1, end_row=hr2, end_column=1)
    _header_cell(ws, hr1, 2, "Perkara"); ws.merge_cells(start_row=hr1, start_column=2, end_row=hr2, end_column=2)

    col_idx = 3
    est_total_excel_col = None
    longtext_excel_cols = []
    supplier_total_positions = {}  # supplier_id -> column index of its "Jumlah" cell
    for c in cols:
        if c["type"] == "supplier-matrix":
            for i, s in enumerate(suppliers):
                _header_cell(ws, hr1, col_idx, f"{i+1}/{len(suppliers)}  {s['name']}")
                ws.merge_cells(start_row=hr1, start_column=col_idx, end_row=hr1, end_column=col_idx + sub_count - 1)
                sc = col_idx
                if fields["unit_price"]:
                    _header_cell(ws, hr2, sc, "Harga/Unit (RM)"); sc += 1
                if fields["total"]:
                    _header_cell(ws, hr2, sc, "Jumlah (RM)")
                    supplier_total_positions[s["id"]] = sc
                    sc += 1
                if fields["status"]:
                    _header_cell(ws, hr2, sc, "Status"); sc += 1
                if sc == col_idx:
                    _header_cell(ws, hr2, sc, ""); sc += 1
                col_idx += sub_count
        else:
            _header_cell(ws, hr1, col_idx, c["label"])
            ws.merge_cells(start_row=hr1, start_column=col_idx, end_row=hr2, end_column=col_idx)
            if c.get("tag") == "estTotal":
                est_total_excel_col = col_idx
            if c["type"] == "longtext":
                longtext_excel_cols.append(col_idx)
            col_idx += 1

    r = hr2 + 1
    qty_col = next((c for c in state["columns"] if c.get("tag") == "quantity"), None)

    for cat in cats:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
        cat_num = S.category_full_number(state, cat["id"])
        cc = ws.cell(row=r, column=1, value=f"{cat_num}. {cat['name']}")
        cc.font = Font(bold=True, size=9)
        cc.fill = PatternFill("solid", fgColor=GOLD)
        r += 1

        items = S.items_for_category(state, cat["id"])
        subtotals = {}
        supplier_subtotals = [{"unit": 0.0, "total": 0.0} for _ in suppliers]

        for idx, item in enumerate(items):
            ws.cell(row=r, column=1, value=f"{cat_num}.{idx+1}").alignment = Alignment(horizontal="center")
            pcell = ws.cell(row=r, column=2, value=item["perkara"])
            pcell.alignment = Alignment(wrap_text=True, vertical="top")

            c = 3
            for col in cols:
                if col["type"] == "supplier-matrix":
                    qty = float(S.get_field(item, qty_col["id"]) or 0) if qty_col else 1
                    for si, s in enumerate(suppliers):
                        bid = S.get_bid(state, item["id"], s["id"])
                        up = bid.get("unit_price")
                        if up is not None:
                            supplier_subtotals[si]["unit"] += up
                            supplier_subtotals[si]["total"] += qty * up
                        st_def = S.status_def(state, bid.get("status"))
                        sc = c
                        if fields["unit_price"]:
                            cell = ws.cell(row=r, column=sc)
                            disp = S.price_display(bid)
                            if up is not None:
                                cell.value = up; cell.number_format = "#,##0.00"
                            elif disp:
                                cell.value = disp
                            sc += 1
                        if fields["total"]:
                            cell = ws.cell(row=r, column=sc)
                            jval, is_num = _jumlah_value(bid, qty)
                            if is_num:
                                cell.value = jval; cell.number_format = "#,##0.00"
                            else:
                                cell.value = jval
                            sc += 1
                        if fields["status"]:
                            cell = ws.cell(row=r, column=sc, value=bid.get("status") or "")
                            cell.alignment = Alignment(horizontal="center")
                            cell.font = Font(bold=True, size=9)
                            if st_def and st_def["color"] in COLOR_HEX:
                                cell.fill = PatternFill("solid", fgColor=COLOR_HEX[st_def["color"]])
                            sc += 1
                        c += sub_count
                else:
                    v = S.compute_column_value(item, col)
                    cell = ws.cell(row=r, column=c)
                    if col["type"] in ("number", "currency", "computed"):
                        if isinstance(v, (int, float)):
                            cell.value = v; cell.number_format = "#,##0.00"
                            subtotals[col["id"]] = subtotals.get(col["id"], 0) + v
                    elif col["type"] == "supplier-pick":
                        cell.value = S.supplier_name(state, v) if v else ""
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.value = v or ""
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                    c += 1
            r += 1

        if items:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            lbl = ws.cell(row=r, column=1, value="Jumlah Kecil")
            lbl.font = Font(bold=True)
            lbl.alignment = Alignment(horizontal="center")
            c = 3
            for col in cols:
                if col["type"] == "supplier-matrix":
                    for si, s in enumerate(suppliers):
                        sc = c
                        if fields["unit_price"]:
                            if supplier_subtotals[si]["unit"]:
                                cell = ws.cell(row=r, column=sc, value=supplier_subtotals[si]["unit"])
                                cell.number_format = "#,##0.00"; cell.font = Font(bold=True)
                            sc += 1
                        if fields["total"]:
                            if supplier_subtotals[si]["total"]:
                                cell = ws.cell(row=r, column=sc, value=supplier_subtotals[si]["total"])
                                cell.number_format = "#,##0.00"; cell.font = Font(bold=True)
                            sc += 1
                        c += sub_count
                else:
                    if col["type"] in ("number", "currency", "computed") and subtotals.get(col["id"]):
                        cell = ws.cell(row=r, column=c, value=subtotals[col["id"]])
                        cell.number_format = "#,##0.00"; cell.font = Font(bold=True)
                    c += 1
            for cc_idx in range(1, total_cols + 1):
                cell = ws.cell(row=r, column=cc_idx)
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = PatternFill("solid", fgColor=SUBTOTAL_BG)
            r += 1

    for row in ws.iter_rows(min_row=1, max_row=r - 1, max_col=total_cols):
        for cell in row:
            cell.border = BORDER
            if cell.font is None or cell.font.size is None:
                cell.font = Font(size=9)

    if cats and scope != "technical":
        r = _write_summary_excel(ws, state, r, total_cols, est_total_excel_col, supplier_total_positions)

    ws.column_dimensions[get_column_letter(2)].width = 42
    for c in range(3, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    for c in longtext_excel_cols:
        ws.column_dimensions[get_column_letter(c)].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


SUMMARY_BG = "FFDCE8F5"
SUMMARY_WARN_BG = "FFFAEADC"


def _write_summary_excel(ws, state, r, total_cols, est_total_col, supplier_total_positions):
    summ = S.compute_summary(state)

    def write_row(label, value=None, value_col=None, value_map=None, bg=SUMMARY_BG, note=None):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        lbl = ws.cell(row=r, column=1, value=label)
        lbl.font = Font(bold=True, size=9)
        for cc in range(1, total_cols + 1):
            ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=bg)
            ws.cell(row=r, column=cc).border = BORDER
        if value is not None and value_col:
            cell = ws.cell(row=r, column=value_col, value=value)
            cell.number_format = "#,##0.00"; cell.font = Font(bold=True, size=9)
        if value_map:
            for sup_id, col_pos in supplier_total_positions.items():
                v = value_map.get(sup_id, 0)
                if v:
                    cell = ws.cell(row=r, column=col_pos, value=v)
                    cell.number_format = "#,##0.00"; cell.font = Font(bold=True, size=9)
        if note:
            ncell = ws.cell(row=r, column=3, value=note)
            ncell.font = Font(italic=True, size=9, color="FFB3541E")
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=total_cols)
        r += 1

    write_row("HARGA ANGGARAN JABATAN (RM)", value=summ["dept_total"], value_col=est_total_col or 3)
    write_row("HARGA TAWARAN DARI PENYEBUT HARGA (RM)", value_map=summ["supplier_totals"])
    if summ["has_pick_column"]:
        write_row("HARGA YANG DICADANGKAN (RM)", value_map=summ["supplier_recommended"])
    else:
        write_row("HARGA YANG DICADANGKAN (RM)", bg=SUMMARY_WARN_BG,
                   note="Tetapkan pilihan Pembekal Dipilih bagi setiap item untuk mengira baris ini.")
    write_row("JUMLAH KESELURUHAN PEROLEHAN (RM)", value=summ["grand_total"], value_col=est_total_col or 3)
    return r
